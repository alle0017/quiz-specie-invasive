import openpyxl
wb = openpyxl.load_workbook('./excel.xlsx',data_only=True)
fs = wb.active
fs_count_row = fs.max_row 
fs_count_col = 4
json = '['
for row in range(1,fs_count_row+1):
    correct = 0
    json += '{ "q": "' + fs.cell(column=1, row=row).value + '", "a": ['
    for column in range(2,fs_count_col+1):
        cell = fs.cell(column=column, row=row)
        json += '"' + cell.value + '"'

        if column != fs_count_col: 
            json += ','
        
    json += '], "correct": "'+ str(correct) +'"}'
    if row != fs_count_row:
        json += ','
json += ']'
file = open("./questions.json", "a")
file.write(json)
file.close()
